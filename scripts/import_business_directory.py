#!/usr/bin/env python3
from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
import os
import sys
from typing import Any, Optional

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


DEFAULT_SOURCES = (
    (
        "/Users/nation/Downloads/pasadena_ca_business_directory_chamber.xlsx",
        "All Chamber Members",
        "pasadena_ca_business_directory_chamber",
    ),
    (
        "/Users/nation/Downloads/arcadia_glendale_burbank_business_directory.xlsx",
        "Directory",
        "arcadia_glendale_burbank_business_directory",
    ),
)

FIELD_ALIASES = {
    "requested_city": ("requested city",),
    "business_name": ("business name", "name"),
    "business_type": ("business type", "category", "industry"),
    "contact_person": ("contact person", "contact"),
    "phone_number": ("phone number", "phone"),
    "fax": ("fax",),
    "address": ("address",),
    "city": ("city",),
    "state": ("state",),
    "zip_code": ("zip", "zip code", "zipcode", "postal code"),
    "email": ("email", "e-mail"),
    "website": ("website", "web site", "url"),
    "source_month_page": ("source month/page", "source month page"),
    "source_url": ("source url",),
    "data_source": ("data source",),
    "city_match": ("city match",),
    "coverage_notes": ("coverage notes",),
    "notes": ("notes",),
    "pdf_page": ("pdf page",),
    "printed_page": ("printed page",),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import PerkNation business directory Excel files.")
    parser.add_argument(
        "--source",
        action="append",
        default=[],
        help="Workbook source as PATH, PATH::SHEET, or PATH::SHEET::LABEL. Defaults to the two attached files.",
    )
    parser.add_argument("--env-file", help="Optional env file for database settings, passed as PERKNATION_ENV_FILE.")
    parser.add_argument("--database-url", help="Optional DATABASE_URL override.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report rows without writing to the DB.")
    parser.add_argument("--skip-create-tables", action="store_true", help="Do not run SQLAlchemy create_all before import.")
    parser.add_argument("--enrich-online", action="store_true", help="Fetch official website metadata for descriptions/media.")
    parser.add_argument("--enrich-limit", type=int, default=0, help="Max website metadata fetches. 0 means no limit.")
    parser.add_argument("--enrich-timeout", type=float, default=4.0, help="Seconds to wait for each website metadata fetch.")
    parser.add_argument("--commit-every", type=int, default=200, help="Commit interval for live imports.")
    parser.add_argument(
        "--dedupe-natural",
        action="store_true",
        help="Update matching active listings by business name plus city/address/phone/website when source rows moved.",
    )
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_source_spec(spec: str) -> tuple[str, Optional[str], str]:
    parts = spec.split("::")
    path = parts[0].strip()
    sheet = parts[1].strip() if len(parts) >= 2 and parts[1].strip() else None
    label = parts[2].strip() if len(parts) >= 3 and parts[2].strip() else Path(path).stem
    return path, sheet, label


def configured_sources(args: argparse.Namespace) -> list[tuple[str, Optional[str], str]]:
    if args.source:
        return [parse_source_spec(source) for source in args.source]
    return [(path, sheet, label) for path, sheet, label in DEFAULT_SOURCES if Path(path).exists()]


def pick_field(raw_by_header: dict[str, str], field: str) -> str:
    for alias in FIELD_ALIASES[field]:
        value = raw_by_header.get(alias, "")
        if value:
            return value
    return ""


def iter_records(path: str, sheet_name: Optional[str]) -> tuple[str, list[tuple[int, dict[str, Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    resolved_sheet = sheet_name or workbook.sheetnames[0]
    if resolved_sheet not in workbook.sheetnames:
        raise ValueError(f"{Path(path).name}: sheet {resolved_sheet!r} not found. Available: {', '.join(workbook.sheetnames)}")

    sheet = workbook[resolved_sheet]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration:
        return resolved_sheet, []

    headers = [normalize_header(value) for value in header_values]
    display_headers = [cell_to_text(value) for value in header_values]
    records: list[tuple[int, dict[str, Any]]] = []
    for row_index, row_values in enumerate(rows, start=2):
        raw_by_header: dict[str, str] = {}
        source_columns: dict[str, str] = {}
        for index, value in enumerate(row_values):
            header = headers[index] if index < len(headers) else ""
            display_header = display_headers[index] if index < len(display_headers) else ""
            if not header:
                continue
            text = cell_to_text(value)
            raw_by_header[header] = text
            if display_header:
                source_columns[display_header] = text

        business_name = pick_field(raw_by_header, "business_name")
        if not business_name:
            continue

        record = {field: pick_field(raw_by_header, field) for field in FIELD_ALIASES}
        record["business_name"] = business_name
        record["source_columns"] = source_columns
        records.append((row_index, record))

    return resolved_sheet, records


def main() -> int:
    args = parse_args()
    if args.env_file:
        os.environ["PERKNATION_ENV_FILE"] = args.env_file
    if args.database_url:
        os.environ["DATABASE_URL"] = args.database_url

    sources = configured_sources(args)
    if not sources:
        print("No sources found. Pass --source PATH::SHEET or place the attached files in /Users/nation/Downloads.", file=sys.stderr)
        return 2

    from app.db.base import Base
    from app.db.session import SessionLocal, engine
    from app.services.business_directory import WebsiteMetadata, fetch_website_metadata, upsert_business_directory_entry

    if not args.skip_create_tables and not args.dry_run:
        Base.metadata.create_all(bind=engine)

    counters: Counter[str] = Counter()
    website_cache: dict[str, WebsiteMetadata] = {}
    enrich_attempts = 0

    with SessionLocal() as db:
        for path, sheet_name, source_label in sources:
            resolved_sheet, records = iter_records(path, sheet_name)
            print(f"{Path(path).name} / {resolved_sheet}: {len(records)} source rows")
            counters["source_rows"] += len(records)

            for row_index, record in records:
                metadata = WebsiteMetadata()
                website = str(record.get("website") or "").strip()
                can_enrich = bool(args.enrich_online and website)
                if can_enrich and (args.enrich_limit <= 0 or enrich_attempts < args.enrich_limit):
                    if website in website_cache:
                        metadata = website_cache[website]
                    else:
                        enrich_attempts += 1
                        try:
                            metadata = fetch_website_metadata(website, timeout_seconds=args.enrich_timeout)
                            counters["metadata_ok"] += 1 if metadata.description or metadata.image_url or metadata.video_url else 0
                        except Exception:
                            counters["metadata_failed"] += 1
                            metadata = WebsiteMetadata()
                        website_cache[website] = metadata

                if args.dry_run:
                    counters["parsed"] += 1
                    continue

                _entry, created = upsert_business_directory_entry(
                    db,
                    source_file=source_label,
                    source_sheet=resolved_sheet,
                    source_row=row_index,
                    raw_record=record,
                    metadata=metadata,
                    dedupe_natural=args.dedupe_natural,
                )
                counters["created" if created else "updated"] += 1
                if (counters["created"] + counters["updated"]) % max(1, args.commit_every) == 0:
                    db.commit()

        if not args.dry_run:
            db.commit()

    print(
        "Import summary: "
        + ", ".join(f"{key}={value}" for key, value in sorted(counters.items()))
        + f", enrich_attempts={enrich_attempts}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
